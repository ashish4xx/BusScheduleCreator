import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font
import base64

# Inject custom CSS to style the Streamlit app
st.markdown(
    """
    <style>
    /* Change background color and font */
    body {
        background-color: #f4f4f4;
        font-family: 'Arial', sans-serif;
    }
    /* Style for the main title */
    .stApp header {
        background-color: #004466;
        color: #ffffff;
        padding: 10px;
        text-align: center;
        font-size: 24px;
        font-weight: bold;
    }
    /* Style for the headers in the app */
    h1, h2, h3 {
        color: #004466;
    }
    /* Style for text input boxes */
    .stTextInput, .stTextArea, .stNumberInput {
        border: 1px solid #004466;
        border-radius: 5px;
        padding: 5px;
    }
    /* Style for the file uploader */
    .stFileUploader {
        border: 1px solid #004466;
        border-radius: 5px;
        padding: 10px;
    }
    /* Style for the buttons */
    .stButton button {
        background-color: #004466;
        color: #ffffff;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
        font-size: 16px;
        cursor: pointer;
    }
    .stButton button:hover {
        background-color: #003355;
    }
    /* Center the images */
    .stImage img {
        display: block;
        margin-left: auto;
        margin-right: auto;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Function to create the Excel file
def create_bus_schedule(routes, stop_id_df, trip_group_name):
    # Define the column headers
    columns = [
        "Trip_group", "route_id", "service_id", "direction_id", "trip_headsign",
        "stop_id", "stop_name", "stop_headsign", "pickup_type", "drop_off_type",
        "continuous_pickup", "continuous_drop_off", "timed_stop", "wait_time",
        "timepoint", "wheelchair_accessible", "bikes_allowed", "T01", "T02", "T03",
        "T04", "T05", "T06", "T07", "T08", "T09", "T10", "T11", "T12", "T13", "T14",
        "T15", "T16", "T17", "T18", "T19", "T20", "T21", "T22", "T23", "T24", "T25",
        "T26", "T27", "T28", "T29", "T30", "T31", "T32"
    ]

    # Define the column headers for the routes sheet
    route_columns = [
        "route_id", "agency_id", "route_short_name", "route_long_name", "route_desc",
        "route_type", "route_url", "route_color", "route_text_color", "route_sort_order",
        "continuous_pickup", "continuous_drop_off"
    ]

    # Create an empty DataFrame with these columns
    df = pd.DataFrame(columns=columns)
    route_df = pd.DataFrame(columns=route_columns)

    # Process each route
    for route_num, (route, time_schedule_zones, ac_type) in enumerate(routes, start=1):
        stop_names, start_end_times = route
        stop_names = [name.strip() for name in stop_names.split('\n') if name.strip()]
        num_stops = len(stop_names)

        start_time_str, end_time_str = start_end_times
        start_time = datetime.strptime(start_time_str, "%H:%M:%S")
        end_time = datetime.strptime(end_time_str, "%H:%M:%S")

        total_time = (end_time - start_time).total_seconds()
        interval_seconds = total_time / (num_stops - 1)
        interval = timedelta(seconds=interval_seconds)

        stops_with_times = [(stop_names[i], (start_time + i * interval).strftime("%H:%M:%S")) for i in range(num_stops)]

        # Get the initials from the first and last stop names and create route_id
        if stops_with_times:
            first_stop_initials = ''.join([word[0] for word in stops_with_times[0][0].split()[:2]]).upper()
            last_stop_initials = ''.join([word[0] for word in stops_with_times[-1][0].split()[:2]]).upper()
            if ac_type == "AC":
                route_id = (first_stop_initials[0] + last_stop_initials[0] + "AC")
                route_long_name = f"{stops_with_times[0][0]} - {stops_with_times[-1][0]} AC"
            else:
                route_id = (first_stop_initials + last_stop_initials)[:4]
                route_long_name = f"{stops_with_times[0][0]} - {stops_with_times[-1][0]}"
        else:
            route_id = "UNKNOWN"
            route_long_name = "UNKNOWN"

        # Add the trip group entry
        df.loc[len(df)] = {"Trip_group": trip_group_name.upper() + str(route_num)}

        # Add the stop names and corresponding stop IDs and times to the DataFrame
        service_id = "FULLW"
        timed_stop = 1
        timepoint = 1

        for i, (stop_name, stop_time) in enumerate(stops_with_times):
            stop_id = stop_id_df.loc[stop_id_df['stop_name'] == stop_name, 'stop_id'].values
            if stop_id.size > 0:
                stop_id_value = stop_id[0]
            else:
                stop_id_value = None  # Leave stop_id blank if stop name is not found

            entry = {
                'route_id': route_id,
                'service_id': service_id,
                'stop_id': stop_id_value,
                'stop_name': stop_name,
                'timed_stop': timed_stop,
                'timepoint': timepoint,
                'T01': stop_time,
            }

            # Initialize current_time with the first trip's stop time
            current_time = datetime.strptime(stop_time, "%H:%M:%S")

            # Initialize trip_index starting from 1, which corresponds to T01
            trip_index = 1

            # Iterate over each time schedule zone provided by the user
            for tz_index, (num_trips, interval) in enumerate(time_schedule_zones):
                # Calculate the interval as a timedelta object
                interval_delta = datetime.strptime(interval, "%H:%M:%S") - datetime(1900, 1, 1)

                # Determine the actual number of trips to calculate for this zone
                if tz_index == 0 and num_trips == 1:
                    trips_to_calculate = num_trips

                elif tz_index == 0:
                    # For the first time schedule zone, calculate (num_trips - 1) trips
                    trips_to_calculate = num_trips - 1

                else:
                    # For the rest of the time schedule zones, calculate the number of trips entered by the user
                    trips_to_calculate = num_trips

                # Generate the specified number of trips for the current time zone
                for _ in range(trips_to_calculate):
                    # Calculate the time for the next trip
                    next_time = (current_time + interval_delta).strftime("%H:%M:%S")

                    # Increment the trip_index to move to the next trip slot (T02, T03, etc.)
                    trip_index += 1

                    # Save the calculated time for the current trip in the corresponding T column
                    entry[f'T{trip_index:02d}'] = next_time

                    # Update current_time to the time of the newly calculated trip
                    current_time += interval_delta

            # Append the entry with all calculated trip times to the DataFrame
            df = pd.concat([df, pd.DataFrame([entry])], ignore_index=True)

        # Add the route details to the route_df DataFrame
        route_entry = {
            "route_id": route_id,
            "agency_id": "apnawahan-mumbai-in",
            "route_short_name": "",
            "route_long_name": route_long_name,
            "route_desc": "",
            "route_type": 3,
            "route_url": "",
            "route_color": "",
            "route_text_color": "",
            "route_sort_order": "",
            "continuous_pickup": "",
            "continuous_drop_off": ""
        }
        route_df = pd.concat([route_df, pd.DataFrame([route_entry])], ignore_index=True)

        # Add the note row
        note_row = pd.Series(
            {"Trip_group": "Note: a row must be skipped between trip_groups. No data can entered in this row."})
        df = pd.concat([df, pd.DataFrame([note_row])], ignore_index=True)

    # Save the DataFrame to a BytesIO object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule', index=False)
        route_df.to_excel(writer, sheet_name='Routes', index=False)

        # Add formatting
        workbook = writer.book
        worksheet = writer.sheets['Schedule']

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for i, row in df.iterrows():
            if "Note:" in str(row["Trip_group"]):
                for col in range(1, len(columns) + 1):  # Excel columns are 1-indexed
                    cell = worksheet.cell(row=i + 2, column=col)
                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
            else:
                for col in range(18, 50):  # Columns T01 to T32 are from 18 to 49 (inclusive)
                    cell = worksheet.cell(row=i + 2, column=col)
                    cell.fill = green_fill

    return output.getvalue()

# Function to convert an image to base64
def get_base64_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()

# Inject custom CSS to change the sidebar background color and width
st.markdown(
    """
    <style>
    /* Change sidebar background color */
    [data-testid="stSidebar"] {
        display: none;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Display the logo at the top of the main content
sidebar_path = "cropped-Site-Icon.jpg"
sidebar_base64 = get_base64_image(sidebar_path)
st.markdown(
    f"""
    <div style="text-align: center;">
        <img src="data:image/jpeg;base64,{sidebar_base64}" alt="Logo" width="100"/>
    </div>
    """,
    unsafe_allow_html=True
)

st.title('Bus Schedule Creator')

# Input fields
trip_group_name = st.text_input('Enter Trip Group Name:')
num_routes = st.number_input('Enter number of routes:', min_value=1, value=1, step=1)

# Display centered image with specific width and caption
image_path = "Stop_Names.png"  # Update with the path to your image file
image_base64 = get_base64_image(image_path)
st.markdown(
    f"""
    <div style="text-align: center;">
        <p><b>Sample Image of Stop names to be entered</b></p>
        <img src="data:image/jpeg;base64,{image_base64}" alt="Bus Schedule Example" width="200" height="300"/>
    </div>
    """,
    unsafe_allow_html=True
)

routes = []
for i in range(num_routes):
    st.header(f'Route {i + 1}')
    st.text(f'Enter all stop names for route {i + 1} (one per line):')
    stop_names = st.text_area(f'Enter all stop names for route {i + 1} (one per line):', label_visibility="collapsed")
    st.text(f'Enter first bus schedule Start time for route {i + 1} (HH:MM:SS):')
    start_time = st.text_input(f'Start time for route {i + 1} (HH:MM:SS):', key=f'start_time_{i}', label_visibility="collapsed")
    st.text(f'Enter first bus schedule End time for route {i + 1} (HH:MM:SS):')
    end_time = st.text_input(f'End time for route {i + 1} (HH:MM:SS):', key=f'end_time_{i}', label_visibility="collapsed")
    ac_type = st.selectbox(f'Is route {i + 1} AC or Non-AC?', ['Non-AC', 'AC'])

    if stop_names and start_time and end_time:
        num_time_zones = st.number_input(f'Enter number of time schedule zones for route {i + 1}', min_value=1, value=1, step=1)

        time_schedule_zones = []
        for tz in range(num_time_zones):
            st.subheader(f'Time Schedule Zone {tz + 1} for Route {i + 1}')
            num_trips = st.number_input(f'Number of trips in zone {tz + 1}:', min_value=1, step=1, key=f'num_trips_{i}_{tz}')
            interval_options = [
                "00:05:00", "00:10:00", "00:15:00", "00:20:00", "00:25:00", "00:30:00",
                "00:35:00", "00:40:00", "00:45:00", "00:50:00", "00:55:00", "01:00:00",
                "01:05:00", "01:10:00", "01:15:00", "01:20:00", "01:25:00", "01:30:00",
                "01:35:00", "01:40:00", "01:45:00", "01:50:00", "01:55:00", "02:00:00"
            ]

            # Use a select box for the interval input
            interval = st.selectbox(
                f'Interval between trips in zone {tz + 1} (HH:MM:SS):',
                interval_options,
                key=f'interval_{i}_{tz}'
            )
            time_schedule_zones.append((num_trips, interval))

        routes.append(((stop_names, (start_time, end_time)), time_schedule_zones, ac_type))

lat_log = "latlog.png"  # Update with the path to your image file
lat_log_base64 = get_base64_image(lat_log)
st.markdown(
    f"""
    <div style="text-align: center;">
        <p><b>Example Excel Sheet For Latitude Longitude Details Required</b></p>
        <img src="data:image/jpeg;base64,{lat_log_base64}" alt="Latitude Longitude Example" width="300"/>
    </div>
    """,
    unsafe_allow_html=True
)

# File uploader
uploaded_file = st.file_uploader("Upload Stop ID Excel file", type=["xlsx"])

if st.button('Create Schedules Excel Sheet'):
    if routes and uploaded_file:
        stop_id_df = pd.read_excel(uploaded_file)
        excel_data = create_bus_schedule(routes, stop_id_df, trip_group_name)
        st.success('Excel file created successfully!')

        # Provide download button
        st.download_button(
            label="Download Excel file",
            data=excel_data,
            file_name="Bus_Schedule_with_Trip_Group_and_Stops.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error('Please enter stop names, start/end times, and time schedule zones for each route, and upload the Stop ID file.')
